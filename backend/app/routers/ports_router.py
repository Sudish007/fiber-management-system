from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional
import io
import openpyxl
from ..database import get_db
from ..models.models import Port, AuditLog
from ..schemas.schemas import PortCreate, PortUpdate, PortResponse, ImportResponse
from ..auth import get_current_user, require_admin

router = APIRouter(prefix="/api/ports", tags=["Ports"])


@router.get("", response_model=list[PortResponse])
def get_ports(
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    query = db.query(Port)
    if status_filter:
        query = query.filter(Port.status == status_filter)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Port.equipment_name.ilike(search_term),
                Port.equipment_ip.ilike(search_term),
                Port.port_number.ilike(search_term),
                Port.fibre_tag.ilike(search_term),
                Port.ddf_name.ilike(search_term),
            )
        )
    return query.order_by(Port.created_at.desc()).all()


@router.get("/export")
def export_ports(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    ports = db.query(Port).order_by(Port.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ports"

    headers = [
        "ID", "Equipment Name", "Equipment IP", "Equipment Type",
        "Port Number", "Port Type", "Fibre Tag", "DDF Name",
        "DDF Port", "Status", "Remarks", "Created At"
    ]
    ws.append(headers)

    for port in ports:
        ws.append([
            port.id, port.equipment_name, port.equipment_ip,
            port.equipment_type, port.port_number, port.port_type,
            port.fibre_tag, port.ddf_name, port.ddf_port,
            port.status, port.remarks,
            port.created_at.strftime("%Y-%m-%d %H:%M") if port.created_at else ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ports_export.xlsx"}
    )


@router.post("/import", response_model=ImportResponse)
async def import_ports(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin)
):
    if not file.filename.endswith(('.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are accepted")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    imported = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for i, row in enumerate(rows, start=2):
        try:
            if not row or not row[0]:
                continue
            # Skip ID column (index 0), start from Equipment Name (index 1)
            col_offset = 1 if len(row) > 10 else 0
            port = Port(
                equipment_name=str(row[col_offset]) if row[col_offset] else "",
                equipment_ip=str(row[col_offset + 1]) if row[col_offset + 1] else "",
                equipment_type=str(row[col_offset + 2]) if row[col_offset + 2] else "",
                port_number=str(row[col_offset + 3]) if row[col_offset + 3] else "",
                port_type=str(row[col_offset + 4]) if row[col_offset + 4] else "",
                fibre_tag=str(row[col_offset + 5]) if len(row) > col_offset + 5 and row[col_offset + 5] else None,
                ddf_name=str(row[col_offset + 6]) if len(row) > col_offset + 6 and row[col_offset + 6] else None,
                ddf_port=str(row[col_offset + 7]) if len(row) > col_offset + 7 and row[col_offset + 7] else None,
                status=str(row[col_offset + 8]) if len(row) > col_offset + 8 and row[col_offset + 8] else "active",
                remarks=str(row[col_offset + 9]) if len(row) > col_offset + 9 and row[col_offset + 9] else None,
            )
            db.add(port)
            imported += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()

    audit = AuditLog(
        user_id=current_user.id,
        action="IMPORT",
        entity_type="port",
        details=f"Imported {imported} ports from {file.filename}"
    )
    db.add(audit)
    db.commit()

    return ImportResponse(imported=imported, errors=errors)


@router.get("/{port_id}", response_model=PortResponse)
def get_port(port_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    port = db.query(Port).filter(Port.id == port_id).first()
    if not port:
        raise HTTPException(status_code=404, detail="Port not found")
    return port


@router.post("", response_model=PortResponse, status_code=status.HTTP_201_CREATED)
def create_port(port: PortCreate, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_port = Port(**port.model_dump())
    db.add(db_port)
    db.commit()
    db.refresh(db_port)

    audit = AuditLog(
        user_id=current_user.id,
        action="CREATE",
        entity_type="port",
        entity_id=db_port.id,
        details=f"Created port {db_port.port_number} on {db_port.equipment_name}"
    )
    db.add(audit)
    db.commit()

    return db_port


@router.put("/{port_id}", response_model=PortResponse)
def update_port(port_id: int, port: PortUpdate, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_port = db.query(Port).filter(Port.id == port_id).first()
    if not db_port:
        raise HTTPException(status_code=404, detail="Port not found")

    update_data = port.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_port, key, value)

    db.commit()
    db.refresh(db_port)

    audit = AuditLog(
        user_id=current_user.id,
        action="UPDATE",
        entity_type="port",
        entity_id=db_port.id,
        details=f"Updated port {db_port.port_number}"
    )
    db.add(audit)
    db.commit()

    return db_port


@router.delete("/{port_id}")
def delete_port(port_id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_port = db.query(Port).filter(Port.id == port_id).first()
    if not db_port:
        raise HTTPException(status_code=404, detail="Port not found")

    audit = AuditLog(
        user_id=current_user.id,
        action="DELETE",
        entity_type="port",
        entity_id=db_port.id,
        details=f"Deleted port {db_port.port_number} from {db_port.equipment_name}"
    )
    db.add(audit)

    db.delete(db_port)
    db.commit()

    return {"message": "Port deleted successfully"}
