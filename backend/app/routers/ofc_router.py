from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_
from typing import Optional
import io
import openpyxl
from ..database import get_db
from ..models.models import OFCRoute, FiberCore, AuditLog
from ..schemas.schemas import (
    OFCRouteCreate, OFCRouteUpdate, OFCRouteResponse,
    FiberCoreCreate, FiberCoreUpdate, FiberCoreResponse, ImportResponse
)
from ..auth import get_current_user, require_admin

router = APIRouter(prefix="/api/ofc", tags=["OFC Routes"])


@router.get("", response_model=list[OFCRouteResponse])
def get_ofc_routes(
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    query = db.query(OFCRoute).options(joinedload(OFCRoute.fiber_cores))
    if status_filter:
        query = query.filter(OFCRoute.status == status_filter)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                OFCRoute.route_name.ilike(search_term),
                OFCRoute.start_location.ilike(search_term),
                OFCRoute.end_location.ilike(search_term),
            )
        )
    return query.order_by(OFCRoute.created_at.desc()).all()


@router.get("/export")
def export_ofc(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    routes = db.query(OFCRoute).options(joinedload(OFCRoute.fiber_cores)).order_by(OFCRoute.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OFC Routes"

    headers = [
        "Route Name", "Start Location", "End Location",
        "Route Length (km)", "Fiber Count", "Core Utilization (%)",
        "Status", "Remarks", "Created At"
    ]
    ws.append(headers)

    for route in routes:
        ws.append([
            route.route_name, route.start_location,
            route.end_location, float(route.route_length) if route.route_length else 0,
            route.fiber_count, route.core_utilization,
            route.status, route.remarks,
            route.created_at.strftime("%Y-%m-%d %H:%M") if route.created_at else ""
        ])

    # Second sheet for fiber cores
    ws2 = wb.create_sheet("Fiber Cores")
    fiber_headers = [
        "Route Name", "Fiber Number", "Color", "Status",
        "From → To", "Connected Equipment", "Port", "Remarks", "Updated At"
    ]
    ws2.append(fiber_headers)

    for route in routes:
        for core in route.fiber_cores:
            ws2.append([
                route.route_name, core.fiber_number, core.color,
                core.status, core.from_to, core.connected_equipment,
                core.port, core.remarks,
                core.updated_at.strftime("%Y-%m-%d %H:%M") if core.updated_at else ""
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ofc_routes_export.xlsx"}
    )


@router.post("/import", response_model=ImportResponse)
async def import_ofc(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin)
):
    if not file.filename.endswith(('.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are accepted")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))

    imported = 0
    errors = []

    # Find the routes sheet (first sheet, or one with "route" in name)
    routes_sheet = None
    fiber_sheet = None
    for name in wb.sheetnames:
        lower_name = name.lower()
        if 'fiber' in lower_name or 'core' in lower_name:
            fiber_sheet = wb[name]
        elif 'route' in lower_name or 'ofc' in lower_name:
            routes_sheet = wb[name]

    # If no specific routes sheet found, use first sheet that's NOT fiber cores
    if not routes_sheet:
        for name in wb.sheetnames:
            lower_name = name.lower()
            if 'fiber' not in lower_name and 'core' not in lower_name:
                routes_sheet = wb[name]
                break

    if not routes_sheet:
        routes_sheet = wb.worksheets[0]

    # Sheet 1: OFC Routes
    rows = list(routes_sheet.iter_rows(min_row=2, values_only=True))

    route_name_map = {}  # map route_name -> route obj for fiber core import

    for i, row in enumerate(rows, start=2):
        try:
            if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            # Find first non-empty cell to detect offset
            values = [str(cell).strip() if cell is not None else '' for cell in row]
            # Skip empty leading columns
            start = 0
            while start < len(values) and values[start] == '':
                start += 1
            if start >= len(values):
                continue

            v = values[start:]
            route_name = v[0] if len(v) > 0 else ""
            if not route_name:
                continue

            # Check if route already exists
            existing = db.query(OFCRoute).filter(OFCRoute.route_name == route_name).first()
            if existing:
                # Update existing
                existing.start_location = v[1] if len(v) > 1 and v[1] else existing.start_location
                existing.end_location = v[2] if len(v) > 2 and v[2] else existing.end_location
                existing.route_length = float(v[3]) if len(v) > 3 and v[3] else existing.route_length
                existing.fiber_count = int(float(v[4])) if len(v) > 4 and v[4] else existing.fiber_count
                existing.core_utilization = int(float(v[5])) if len(v) > 5 and v[5] else existing.core_utilization
                existing.status = v[6] if len(v) > 6 and v[6] else existing.status
                existing.remarks = v[7] if len(v) > 7 and v[7] else existing.remarks
                route_name_map[route_name] = existing
            else:
                route = OFCRoute(
                    route_name=route_name,
                    start_location=v[1] if len(v) > 1 else "",
                    end_location=v[2] if len(v) > 2 else "",
                    route_length=float(v[3]) if len(v) > 3 and v[3] else None,
                    fiber_count=int(float(v[4])) if len(v) > 4 and v[4] else None,
                    core_utilization=int(float(v[5])) if len(v) > 5 and v[5] else None,
                    status=v[6] if len(v) > 6 and v[6] else "active",
                    remarks=v[7] if len(v) > 7 and v[7] else None,
                )
                db.add(route)
                db.flush()
                route_name_map[route_name] = route
            imported += 1
        except Exception as e:
            errors.append(f"Routes Row {i}: {str(e)}")

    # Commit routes first so fiber cores can reference them
    db.commit()

    # Refresh route map with committed IDs
    for name in list(route_name_map.keys()):
        route_obj = db.query(OFCRoute).filter(OFCRoute.route_name == name).first()
        if route_obj:
            route_name_map[name] = route_obj

    # Sheet 2: Fiber Cores (if exists)
    if fiber_sheet:
        fiber_rows = list(fiber_sheet.iter_rows(min_row=2, values_only=True))

        for i, row in enumerate(fiber_rows, start=2):
            try:
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                values = [str(cell).strip() if cell is not None else '' for cell in row]
                start = 0
                while start < len(values) and values[start] == '':
                    start += 1
                if start >= len(values):
                    continue
                v = values[start:]

                route_name = v[0] if len(v) > 0 else ""
                if not route_name:
                    continue

                # Find the route - first from map, then from DB
                route = route_name_map.get(route_name)
                if not route:
                    route = db.query(OFCRoute).filter(OFCRoute.route_name == route_name).first()
                if not route:
                    errors.append(f"Fiber Row {i}: Route '{route_name}' not found")
                    continue

                fiber_number = int(float(v[1])) if len(v) > 1 and v[1] else 1
                color = v[2] if len(v) > 2 and v[2] else "Blue"
                fiber_status = v[3] if len(v) > 3 and v[3] else "spare"
                from_to = v[4] if len(v) > 4 and v[4] else None
                connected_equipment = v[5] if len(v) > 5 and v[5] else None
                port_val = v[6] if len(v) > 6 and v[6] else None
                remarks = v[7] if len(v) > 7 and v[7] else None

                # Check if fiber already exists for this route + number
                existing_fiber = db.query(FiberCore).filter(
                    FiberCore.route_id == route.id,
                    FiberCore.fiber_number == fiber_number
                ).first()

                if existing_fiber:
                    existing_fiber.color = color
                    existing_fiber.status = fiber_status
                    existing_fiber.from_to = from_to
                    existing_fiber.connected_equipment = connected_equipment
                    existing_fiber.port = port_val
                    existing_fiber.remarks = remarks
                else:
                    db_fiber = FiberCore(
                        route_id=route.id,
                        fiber_number=fiber_number,
                        color=color,
                        status=fiber_status,
                        from_to=from_to,
                        connected_equipment=connected_equipment,
                        port=port_val,
                        remarks=remarks,
                    )
                    db.add(db_fiber)
                imported += 1
            except Exception as e:
                errors.append(f"Fiber Row {i}: {str(e)}")

    db.commit()

    # Update utilization for all affected routes
    for route in route_name_map.values():
        used_count = db.query(FiberCore).filter(FiberCore.route_id == route.id, FiberCore.status == "used").count()
        total_count = db.query(FiberCore).filter(FiberCore.route_id == route.id).count()
        route.core_utilization = int((used_count / total_count) * 100) if total_count > 0 else 0
        route.fiber_count = total_count
    db.commit()

    audit = AuditLog(
        user_id=current_user.id,
        action="IMPORT",
        entity_type="ofc_route",
        details=f"Imported {imported} records from {file.filename}"
    )
    db.add(audit)
    db.commit()

    return ImportResponse(imported=imported, errors=errors)


@router.get("/{ofc_id}", response_model=OFCRouteResponse)
def get_ofc_route(ofc_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    route = db.query(OFCRoute).options(joinedload(OFCRoute.fiber_cores)).filter(OFCRoute.id == ofc_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="OFC Route not found")
    return route


@router.post("", response_model=OFCRouteResponse, status_code=status.HTTP_201_CREATED)
def create_ofc_route(route: OFCRouteCreate, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_route = OFCRoute(**route.model_dump())
    db.add(db_route)
    db.commit()
    db.refresh(db_route)

    audit = AuditLog(
        user_id=current_user.id,
        action="CREATE",
        entity_type="ofc_route",
        entity_id=db_route.id,
        details=f"Created OFC route {db_route.route_name}"
    )
    db.add(audit)
    db.commit()

    return db_route


@router.put("/{ofc_id}", response_model=OFCRouteResponse)
def update_ofc_route(ofc_id: int, route: OFCRouteUpdate, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_route = db.query(OFCRoute).filter(OFCRoute.id == ofc_id).first()
    if not db_route:
        raise HTTPException(status_code=404, detail="OFC Route not found")

    update_data = route.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_route, key, value)

    db.commit()
    db.refresh(db_route)

    audit = AuditLog(
        user_id=current_user.id,
        action="UPDATE",
        entity_type="ofc_route",
        entity_id=db_route.id,
        details=f"Updated OFC route {db_route.route_name}"
    )
    db.add(audit)
    db.commit()

    return db_route


@router.delete("/{ofc_id}")
def delete_ofc_route(ofc_id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_route = db.query(OFCRoute).filter(OFCRoute.id == ofc_id).first()
    if not db_route:
        raise HTTPException(status_code=404, detail="OFC Route not found")

    audit = AuditLog(
        user_id=current_user.id,
        action="DELETE",
        entity_type="ofc_route",
        entity_id=db_route.id,
        details=f"Deleted OFC route {db_route.route_name}"
    )
    db.add(audit)

    db.delete(db_route)
    db.commit()

    return {"message": "OFC Route deleted successfully"}


# ============ Fiber Core Endpoints ============

@router.get("/{ofc_id}/fibers", response_model=list[FiberCoreResponse])
def get_fiber_cores(ofc_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    route = db.query(OFCRoute).filter(OFCRoute.id == ofc_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="OFC Route not found")
    return db.query(FiberCore).filter(FiberCore.route_id == ofc_id).order_by(FiberCore.fiber_number).all()


@router.post("/{ofc_id}/fibers", response_model=FiberCoreResponse, status_code=status.HTTP_201_CREATED)
def create_fiber_core(ofc_id: int, fiber: FiberCoreCreate, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    route = db.query(OFCRoute).filter(OFCRoute.id == ofc_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="OFC Route not found")

    db_fiber = FiberCore(route_id=ofc_id, **fiber.model_dump())
    db.add(db_fiber)
    db.commit()
    db.refresh(db_fiber)

    # Update route utilization
    used_count = db.query(FiberCore).filter(FiberCore.route_id == ofc_id, FiberCore.status == "used").count()
    total_count = db.query(FiberCore).filter(FiberCore.route_id == ofc_id).count()
    route.core_utilization = int((used_count / total_count) * 100) if total_count > 0 else 0
    route.fiber_count = total_count
    db.commit()

    audit = AuditLog(
        user_id=current_user.id,
        action="CREATE",
        entity_type="fiber_core",
        entity_id=db_fiber.id,
        details=f"Added fiber #{fiber.fiber_number} ({fiber.color}) to route {route.route_name}"
    )
    db.add(audit)
    db.commit()

    return db_fiber


@router.put("/{ofc_id}/fibers/{fiber_id}", response_model=FiberCoreResponse)
def update_fiber_core(ofc_id: int, fiber_id: int, fiber: FiberCoreUpdate, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_fiber = db.query(FiberCore).filter(FiberCore.id == fiber_id, FiberCore.route_id == ofc_id).first()
    if not db_fiber:
        raise HTTPException(status_code=404, detail="Fiber core not found")

    update_data = fiber.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_fiber, key, value)

    db.commit()
    db.refresh(db_fiber)

    # Update route utilization
    route = db.query(OFCRoute).filter(OFCRoute.id == ofc_id).first()
    used_count = db.query(FiberCore).filter(FiberCore.route_id == ofc_id, FiberCore.status == "used").count()
    total_count = db.query(FiberCore).filter(FiberCore.route_id == ofc_id).count()
    route.core_utilization = int((used_count / total_count) * 100) if total_count > 0 else 0
    db.commit()

    audit = AuditLog(
        user_id=current_user.id,
        action="UPDATE",
        entity_type="fiber_core",
        entity_id=db_fiber.id,
        details=f"Updated fiber #{db_fiber.fiber_number} on route {route.route_name}"
    )
    db.add(audit)
    db.commit()

    return db_fiber


@router.delete("/{ofc_id}/fibers/{fiber_id}")
def delete_fiber_core(ofc_id: int, fiber_id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_fiber = db.query(FiberCore).filter(FiberCore.id == fiber_id, FiberCore.route_id == ofc_id).first()
    if not db_fiber:
        raise HTTPException(status_code=404, detail="Fiber core not found")

    route = db.query(OFCRoute).filter(OFCRoute.id == ofc_id).first()

    audit = AuditLog(
        user_id=current_user.id,
        action="DELETE",
        entity_type="fiber_core",
        entity_id=db_fiber.id,
        details=f"Deleted fiber #{db_fiber.fiber_number} from route {route.route_name}"
    )
    db.add(audit)

    db.delete(db_fiber)
    db.commit()

    # Update route utilization
    used_count = db.query(FiberCore).filter(FiberCore.route_id == ofc_id, FiberCore.status == "used").count()
    total_count = db.query(FiberCore).filter(FiberCore.route_id == ofc_id).count()
    route.core_utilization = int((used_count / total_count) * 100) if total_count > 0 else 0
    route.fiber_count = total_count
    db.commit()

    return {"message": "Fiber core deleted successfully"}
