from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import Optional
import io
import openpyxl
from ..database import get_db
from ..models.models import DDFLog, AuditLog
from ..schemas.schemas import DDFCreate, DDFUpdate, DDFResponse, ImportResponse
from ..auth import get_current_user, require_admin

router = APIRouter(prefix="/api/ddf", tags=["DDF Management"])


@router.get("", response_model=list[DDFResponse])
def get_ddf_records(
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    query = db.query(DDFLog)
    if status_filter:
        query = query.filter(DDFLog.status == status_filter)
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                DDFLog.ddf_name.ilike(search_term),
                DDFLog.ddf_port.ilike(search_term),
                DDFLog.connected_to.ilike(search_term),
            )
        )
    return query.order_by(DDFLog.created_at.desc()).all()


@router.get("/export")
def export_ddf(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    records = db.query(DDFLog).order_by(DDFLog.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DDF Records"

    headers = ["ID", "DDF Name", "DDF Port", "Connected To", "Connection Type", "Status", "Remarks", "Created At"]
    ws.append(headers)

    for record in records:
        ws.append([
            record.id, record.ddf_name, record.ddf_port,
            record.connected_to, record.connection_type,
            record.status, record.remarks,
            record.created_at.strftime("%Y-%m-%d %H:%M") if record.created_at else ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ddf_export.xlsx"}
    )


@router.post("/import", response_model=ImportResponse)
async def import_ddf(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin)
):
    if not file.filename.endswith(('.xlsx', '.csv')):
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are accepted")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    imported = 0
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for i, row in enumerate(rows, start=2):
        try:
            if not row or not row[0]:
                continue
            # Skip ID column if present
            col_offset = 1 if len(row) > 7 else 0
            record = DDFLog(
                ddf_name=str(row[col_offset]) if row[col_offset] else "",
                ddf_port=str(row[col_offset + 1]) if row[col_offset + 1] else "",
                connected_to=str(row[col_offset + 2]) if len(row) > col_offset + 2 and row[col_offset + 2] else None,
                connection_type=str(row[col_offset + 3]) if len(row) > col_offset + 3 and row[col_offset + 3] else None,
                status=str(row[col_offset + 4]) if len(row) > col_offset + 4 and row[col_offset + 4] else "active",
                remarks=str(row[col_offset + 5]) if len(row) > col_offset + 5 and row[col_offset + 5] else None,
            )
            db.add(record)
            imported += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()

    audit = AuditLog(
        user_id=current_user.id,
        action="IMPORT",
        entity_type="ddf",
        details=f"Imported {imported} DDF records from {file.filename}"
    )
    db.add(audit)
    db.commit()

    return ImportResponse(imported=imported, errors=errors)


@router.get("/{ddf_id}", response_model=DDFResponse)
def get_ddf(ddf_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    record = db.query(DDFLog).filter(DDFLog.id == ddf_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="DDF record not found")
    return record


@router.post("", response_model=DDFResponse, status_code=status.HTTP_201_CREATED)
def create_ddf(ddf: DDFCreate, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_ddf = DDFLog(**ddf.model_dump())
    db.add(db_ddf)
    db.commit()
    db.refresh(db_ddf)

    audit = AuditLog(
        user_id=current_user.id,
        action="CREATE",
        entity_type="ddf",
        entity_id=db_ddf.id,
        details=f"Created DDF record {db_ddf.ddf_name}:{db_ddf.ddf_port}"
    )
    db.add(audit)
    db.commit()

    return db_ddf


@router.put("/{ddf_id}", response_model=DDFResponse)
def update_ddf(ddf_id: int, ddf: DDFUpdate, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_ddf = db.query(DDFLog).filter(DDFLog.id == ddf_id).first()
    if not db_ddf:
        raise HTTPException(status_code=404, detail="DDF record not found")

    update_data = ddf.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_ddf, key, value)

    db.commit()
    db.refresh(db_ddf)

    audit = AuditLog(
        user_id=current_user.id,
        action="UPDATE",
        entity_type="ddf",
        entity_id=db_ddf.id,
        details=f"Updated DDF record {db_ddf.ddf_name}:{db_ddf.ddf_port}"
    )
    db.add(audit)
    db.commit()

    return db_ddf


@router.delete("/{ddf_id}")
def delete_ddf(ddf_id: int, db: Session = Depends(get_db), current_user=Depends(require_admin)):
    db_ddf = db.query(DDFLog).filter(DDFLog.id == ddf_id).first()
    if not db_ddf:
        raise HTTPException(status_code=404, detail="DDF record not found")

    audit = AuditLog(
        user_id=current_user.id,
        action="DELETE",
        entity_type="ddf",
        entity_id=db_ddf.id,
        details=f"Deleted DDF record {db_ddf.ddf_name}:{db_ddf.ddf_port}"
    )
    db.add(audit)

    db.delete(db_ddf)
    db.commit()

    return {"message": "DDF record deleted successfully"}
